import openpyxl


wb = openpyxl.load_workbook('example.xlsx')
print(wb.sheetnames)
sheet = wb['Sheet1']
print(sheet['A1'].value)
c = sheet['B1']
print('Row' + str(c.row) +', Column' + c.column + ' is ' + c.value)
for i in range(1, 8, 2):
    print(i,sheet.cell(row = i,column = 2).value)
print(sheet.max_row)
